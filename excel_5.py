import os
import openpyxl
import telebot
import shutil

bot = telebot.TeleBot('1937067149:AAGnI-FHNXK_9BYYJBUYWqKWnBHXp__l97I')
chatID = '@GEO_Dudocknik'

list_sheet = ['ЛЧ_траншея', 'ЛЧ_подушка', 'ЛЧ_укладка', 'ЛЧ_обсыпка']

report_file = 'test.xlsx'
last_report_file = 'write.xlsx'
data_time_file = 'write.xlsx'

book_report = {}
book_last = {}


def parsing():
    for k in range(0, len(list_sheet)):
        book_report.clear()
        book_last.clear()

        def append_in_list(count, dicts, open_file):
            wb = openpyxl.reader.excel.load_workbook(filename=open_file)
            sheet = wb[list_sheet[k]]
            count_row_in_sheet = sheet.max_row + 3

            for num in range(count_row_in_sheet):
                keys = sheet[f'B{count}'].value
                if sheet[f'B{count}'].value is not None:
                    if str(sheet[f'C{count}'].value).find('00:00:00') > 0:
                        data = str(sheet[f'C{count}'].value)[:10].split('-')
                        data_out = f'{data[2]}.{data[1]}.{data[0]}'
                        dicts[f'{keys}'] = [data_out, float(sheet[f'D{count}'].value), float(sheet[f'E{count}'].value),
                                            (float(sheet[f'E{count}'].value) - float(sheet[f'D{count}'].value))]
                    else:
                        dicts[f'{keys}'] = [str(sheet[f'C{count}'].value),
                                            float(sheet[f'D{count}'].value),
                                            float(sheet[f'E{count}'].value),
                                            (float(sheet[f'E{count}'].value) - float(sheet[f'D{count}'].value))]
                count += 1

        def write_new():
            wb_1 = openpyxl.reader.excel.load_workbook(filename=last_report_file)
            new_sheet = wb_1[list_sheet[k]]
            number = list(book_report.keys())
            data = list(book_report.values())

            def write_sheet():
                for i in range(len(number)):
                    c_number = new_sheet[f'B{i + 1}']
                    c_number.value = f'{number[i]}'
                    for j in range(len(data[i])):
                        c_data = new_sheet[f'C{i + 1}']
                        c_data.value = f'{data[i][0]}'

                        c_pk_start = new_sheet[f'D{i + 1}']
                        c_pk_start.value = f'{data[i][1]}'

                        c_pk_end = new_sheet[f'E{i + 1}']
                        c_pk_end.value = f'{data[i][2]}'

                        c_dist = new_sheet[f'F{i + 1}']
                        c_dist.value = f'{data[i][3]}'

            write_sheet()
            wb_1.save(last_report_file)

        def analysis_row_in_report(row_last, row_new):
            L = []
            for index in range(0, len(row_last)):
                if row_last[index] != row_new[index]:
                    if index == 0:
                        L.append(f'дату с {row_last[index]} на {row_new[index]}')
                    if index == 1 or index == 2:
                        pk_index_1_old = str(row_last[1])
                        pk_index_2_old = str(row_last[2])
                        len_1_old = len(pk_index_1_old) - 4
                        len_2_old = len(pk_index_2_old) - 4
                        pk_old = \
                            f'ПК{pk_index_1_old[:len_1_old]}+' \
                            f'{pk_index_1_old[len_1_old:]} - ' \
                            f'ПК{pk_index_2_old[:len_2_old]}+' \
                            f'{pk_index_2_old[len_2_old:]}'

                        pk_index_1_new = str(row_new[1])
                        pk_index_2_new = str(row_new[2])
                        len_1_new = len(pk_index_1_new) - 4
                        len_2_new = len(pk_index_2_new) - 4
                        pk_new = f'ПК{pk_index_1_new[:len_1_new]}+' \
                                 f'{pk_index_1_new[len_1_new:]} - ' \
                                 f'ПК{pk_index_2_new[:len_2_new]}+' \
                                 f'{pk_index_2_new[len_2_new:]}'
                        #  L.append(f'пикеты с {row_last[1]}-{row_last[2]} на {row_new[1]}-{row_new[2]}')
                        L.append(f'пикеты с {pk_old} на {pk_new}')
            if len(L) == 1:
                return f'\n{L[0]}'
            else:
                return f'\n- {L[0]},\n- {L[1]}'

        def overwriting_report(key, list_new_row):
            wb_overwriting = openpyxl.reader.excel.load_workbook(filename=last_report_file)
            new_sheet = wb_overwriting[list_sheet[k]]
            count_row = new_sheet.max_row + 2
            number = list(book_report.keys())
            #  print(f'count_row: {count_row}   key: {key}   list_new_row: {list_new_row}')
            for row in range(count_row):
                #  print(number[row], type(number[row]), type(key), row)
                if key == number[row]:
                    print(number[row])
                    c_data = new_sheet[f'C{row + 1}']
                    c_data.value = f'{list_new_row[0]}'

                    c_pk_start = new_sheet[f'D{row + 1}']
                    c_pk_start.value = f'{list_new_row[1]}'

                    c_pk_end = new_sheet[f'E{row + 1}']
                    c_pk_end.value = f'{list_new_row[2]}'

                    c_dist = new_sheet[f'F{row + 1}']
                    var_dist = list_new_row[2] - list_new_row[1]
                    c_dist.value = f'{var_dist}'
            wb_overwriting.save(last_report_file)

        def analysis_report(dict_last, dict_new):
            for key_report in dict_new:
                try:
                    if dict_new[key_report] != dict_last[key_report]:
                        print(f'В схеме {key_report} изменили: {analysis_row_in_report(dict_last[key_report], dict_new[key_report])}.')
                        #bot.send_message(chatID,
                                         #text=f'В схеме {key_report} изменили: '
                                              #f'{analysis_row_in_report(dict_last[key_report], dict_new[key_report])}.')
                        overwriting_report(key_report, dict_new[key_report])

                except KeyError:
                    pass
                    # print(key_report)

        append_in_list(3, book_report, report_file)
        append_in_list(1, book_last, last_report_file)

        analysis_report(book_last, book_report)

        margin = list(set(book_report) - set(book_last))
        if len(margin) > 0:
            write_new()
            str_info = ''
            # print(f'Добавлены следующие схемы в {list_sheet[k]}:')
            bot.send_message(chatID, text=f'Добавлены следующие схемы в {list_sheet[k]}:')
            for g in range(len(margin)):
                add_pos = book_report[margin[g]]
                pk_1_split = str(add_pos[1]).split('.')
                pk_2_split = str(add_pos[2]).split('.')
                str_info += \
                    f'{margin[g]} {add_pos[0]} ' \
                    f'ПК{pk_1_split[0][:len(pk_1_split[0]) - 2]}+{pk_1_split[0][len(pk_1_split[0]) - 2:]} - ' \
                    f'ПК{pk_2_split[0][:len(pk_1_split[0]) - 2]}+{pk_2_split[0][len(pk_1_split[0]) - 2:]} ' \
                    f'L={round(add_pos[3], 2)} м\n\n'
            bot.send_message(chatID, text=str_info)
            # print(str_info)
        '''    
        else:
            print(f'Добавлений в {list_sheet[k]} не производилось!')
            bot.send_message(chatID, text=f'Добавлений в {list_sheet[k]} не производилось!')
        '''


def check_data_time():
    if os.path.exists("C:\\parsing_report\\test.xlsx"):
        os.remove("C:\\parsing_report\\test.xlsx")
    shutil.copyfile("\\\\Nas\\pps-11\\Геодезия\\Сводка по ИС ППС11 СМУ 11.2 Окунайский.xlsx",
                    "C:\\parsing_report\\test.xlsx")
    wb_data_time = openpyxl.reader.excel.load_workbook(filename='data_time.xlsx')
    sheet_data_time = wb_data_time['data_time']
    data_time_last = float(sheet_data_time['A1'].value)
    data_time = os.stat("\\\\Nas\\pps-11\\Геодезия\\Сводка по ИС ППС11 СМУ 11.2 Окунайский.xlsx").st_mtime
    if data_time_last < data_time:
        parsing()
        write_data_time = sheet_data_time['A1']
        write_data_time.value = str(os.stat('test.xlsx').st_mtime)

        wb_data_time.save(filename='data_time.xlsx')
        os.remove("C:\\parsing_report\\test.xlsx")
    else:
        os.remove("C:\\parsing_report\\test.xlsx")


check_data_time()
#  parsing()
